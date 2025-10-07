import streamlit as st
import pandas as pd
import os
import glob
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import tempfile
import io

# =====================================================================================
# FUNÇÃO PARA O FLUXO "PRODUTO ACABADO" (3 ETAPAS)
# =====================================================================================
def processar_fluxo_produto_acabado(pasta_csv_origem, pasta_trabalho_temporaria):
    """
    Executa o processo completo de 3 etapas para "Produto Acabado".
    """
    
    # --- ETAPA 1: Converter CSV para Excel individual formatado ---
    st.write("Etapa 1 de 3: Convertendo cada CSV para um Excel individual...")
    pasta_excel_intermediario = os.path.join(pasta_trabalho_temporaria, "excel_intermediario")
    os.makedirs(pasta_excel_intermediario, exist_ok=True)

    arquivos_csv = glob.glob(os.path.join(pasta_csv_origem, '*.csv'))
    if not arquivos_csv:
        st.error("Nenhum arquivo .csv foi encontrado para processar.")
        return None

    for caminho_csv in arquivos_csv:
        df = pd.read_csv(caminho_csv, header=None)
        codigos = df[4].str.split(" -", n=1, expand=True)
        codigo1 = codigos[0].str.split("-", expand=True)
        codigo2 = codigos[1].str.split("-", expand=True)
        df_final = pd.concat([df.iloc[:, :4], codigo1, codigo2, df.iloc[:, 5:]], axis=1)
        df_final.columns = [
            "DT Leitura", "HR Leitura", "Reg", "Leitor", "Filial", "Código",
            "Armazem", "Lote", "Peso", "SI", "NV DT", "NV HR", "Coluna 8", "Coluna 9"
        ]
        df_final["Peso"] = pd.to_numeric(df_final["Peso"], errors='coerce') / 1000
        df_final["Peso"] = df_final["Peso"].round(3)

        nome_base = os.path.splitext(os.path.basename(caminho_csv))[0]
        caminho_excel = os.path.join(pasta_excel_intermediario, f"{nome_base}.xlsx")
        
        with pd.ExcelWriter(caminho_excel, engine='openpyxl') as writer:
            df_final.to_excel(writer, sheet_name=nome_base[:31], index=False)
        
        wb = load_workbook(caminho_excel)
        ws = wb.active
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 2
        wb.save(caminho_excel)

    st.success("Etapa 1 concluída!")

    # --- ETAPA 2: Unificar Excels individuais em um único arquivo com múltiplas abas ---
    st.write("Etapa 2 de 3: Unificando arquivos em um único Excel com múltiplas abas...")
    arquivo_unificado_path = os.path.join(pasta_trabalho_temporaria, "Inventario.xlsx")
    wb_final = Workbook()
    wb_final.remove(wb_final.active)

    for nome_arquivo in os.listdir(pasta_excel_intermediario):
        if nome_arquivo.lower().endswith(".xlsx"):
            caminho_arquivo = os.path.join(pasta_excel_intermediario, nome_arquivo)
            wb_origem = load_workbook(caminho_arquivo)
            aba_origem = wb_origem.active
            nome_aba = os.path.splitext(nome_arquivo)[0][:31]
            aba_destino = wb_final.create_sheet(title=nome_aba)

            for row in aba_origem.iter_rows():
                for cell in row:
                    nova_celula = aba_destino.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                    if cell.has_style:
                        nova_celula.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color)
                        nova_celula.fill = PatternFill(fill_type=cell.fill.fill_type, start_color=cell.fill.start_color, end_color=cell.fill.end_color)
                        nova_celula.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=cell.alignment.wrap_text)
            
            for col in aba_destino.columns:
                max_length = max(len(str(cell.value)) for cell in col if cell.value)
                aba_destino.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    wb_final.save(arquivo_unificado_path)
    st.success("Etapa 2 concluída!")
    
    # --- ETAPA 3: Consolidar todas as abas em uma única aba final ---
    st.write("Etapa 3 de 3: Consolidando todas as abas em um relatório final...")
    dfs_coletados = []
    cabecalho_final = None

    excel_file = pd.ExcelFile(arquivo_unificado_path)
    workbook_leitura = load_workbook(arquivo_unificado_path, read_only=True)
    
    for nome_da_aba in excel_file.sheet_names:
        df_cabecalho_temp = pd.read_excel(excel_file, sheet_name=nome_da_aba, header=0, nrows=0, dtype=str)
        nomes_colunas_da_aba = df_cabecalho_temp.columns.tolist()
        if cabecalho_final is None: cabecalho_final = nomes_colunas_da_aba

        worksheet_ativa = workbook_leitura[nome_da_aba]
        valor_celula_a2 = worksheet_ativa['A2'].value
        linhas_a_pular = 1
        if valor_celula_a2 and "date" in str(valor_celula_a2).lower():
            linhas_a_pular = 2

        df_dados = pd.read_excel(excel_file, sheet_name=nome_da_aba, header=None, skiprows=linhas_a_pular, dtype=str)
        
        if not df_dados.empty:
            df_dados.columns = nomes_colunas_da_aba[:len(df_dados.columns)]
            df_dados["Localizacao"] = nome_da_aba
            dfs_coletados.append(df_dados)

    if not dfs_coletados:
        st.error("Nenhum dado encontrado nas abas para consolidar.")
        return None

    df_final_consolidado = pd.concat(dfs_coletados, ignore_index=True)
    
    # Aplicar formatações específicas do script converter_inventariov2.py
    if "Armazem" in df_final_consolidado.columns:
        df_final_consolidado["Armazem"] = df_final_consolidado["Armazem"].fillna('').apply(lambda val: f"{int(float(str(val).strip())):02d}" if str(val).strip().replace('.','',1).isdigit() else str(val).strip())
    if "Lote" in df_final_consolidado.columns:
        df_final_consolidado["Lote"] = df_final_consolidado["Lote"].fillna('').astype(str)
    if "Peso" in df_final_consolidado.columns:
        df_final_consolidado["Peso"] = pd.to_numeric(df_final_consolidado["Peso"].str.replace(',', '.', 1), errors='coerce')

    arquivo_final_path = os.path.join(pasta_trabalho_temporaria, "Inventario_Final.xlsx")
    with pd.ExcelWriter(arquivo_final_path, engine='openpyxl') as writer:
        df_final_consolidado.to_excel(writer, index=False, sheet_name="Inventario Geral")
    
    wb = load_workbook(arquivo_final_path)
    ws = wb.active
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
    if "Peso" in df_final_consolidado.columns:
        col_idx = df_final_consolidado.columns.get_loc("Peso") + 1
        for cell in ws[get_column_letter(col_idx)]:
            if isinstance(cell.value, (int, float)): cell.number_format = '0.000'
    wb.save(arquivo_final_path)

    st.success("Etapa 3 concluída!")
    return arquivo_final_path

# =====================================================================================
# FUNÇÃO PARA O FLUXO "BOBINA" (2 ETAPAS) - Sem alterações
# =====================================================================================
def processar_fluxo_bobina(pasta_csv_origem, pasta_trabalho_temporaria):
    st.write("Etapa 1 de 2: Convertendo cada CSV para um Excel formatado...")
    arquivos_csv = glob.glob(os.path.join(pasta_csv_origem, '*.csv'))
    if not arquivos_csv:
        st.error("Nenhum arquivo .csv foi encontrado para processar.")
        return None

    for file_path in arquivos_csv:
        dados_processados = []
        try:
            try: df_csv = pd.read_csv(file_path, header=None, encoding='utf-8', low_memory=False)
            except UnicodeDecodeError: df_csv = pd.read_csv(file_path, header=None, encoding='latin1', low_memory=False)
            mask = df_csv.apply(lambda row: 'date' not in ' '.join(row.astype(str)).lower(), axis=1)
            df_csv = df_csv[mask]
        except Exception as e:
            st.warning(f"Não foi possível ler o arquivo {os.path.basename(file_path)}. Erro: {e}")
            continue

        for index, row in df_csv.iterrows():
            if len(row) < 5: continue
            data_leitura, hora_leitura, _, tipo_codigo, dados_lidos = row[0:5]
            dados_lidos, tipo_codigo = str(dados_lidos).strip(), str(tipo_codigo).strip()
            nova_linha = {"Data da Leitura": data_leitura, "Hora da Leitura": hora_leitura, "Lote": None, "Peso": None}
            if tipo_codigo == 'Code128':
                if ' ' in dados_lidos: nova_linha.update({"Lote": "erro de leitura", "Peso": "erro de leitura"})
                elif '*' in dados_lidos:
                    try:
                        partes = dados_lidos.split('*')
                        if dados_lidos.startswith('*'): nova_linha.update({"Lote": partes[3].strip(), "Peso": float(partes[2].strip()) / 1000.0})
                        else: nova_linha.update({"Lote": partes[2].strip(), "Peso": float(partes[1].strip()) / 1000.0})
                    except (ValueError, IndexError): nova_linha.update({"Lote": "erro Code128/*", "Peso": "erro Code128/*"})
                else:
                    if dados_lidos.isdigit() and len(dados_lidos) <= 5: nova_linha.update({"Peso": float(dados_lidos) / 1000.0, "Lote": None})
                    else: nova_linha.update({"Lote": dados_lidos, "Peso": None})
            elif tipo_codigo in ['CODE_39', 'CODE_128']:
                nova_linha["Data da Leitura"] = datetime.strptime(str(data_leitura), '%m-%d-%Y').strftime('%d/%m/%Y')
                nova_linha.update({"Lote": dados_lidos, "Peso": None})
            elif tipo_codigo in ['QR_CODE', 'QR']:
                nova_linha["Data da Leitura"] = datetime.strptime(str(data_leitura), '%m-%d-%Y').strftime('%d/%m/%Y')
                if '{' in dados_lidos and '}' in dados_lidos:
                    try:
                        partes = dados_lidos.split('{', 1)
                        identificador = partes[0].strip('"-')
                        dados_json = json.loads('{' + partes[1])
                        nova_linha.update({"Peso": float(dados_json.get('peso', 0)), "Lote": identificador})
                    except (ValueError, IndexError, json.JSONDecodeError): nova_linha.update({"Lote": "erro QR/JSON", "Peso": "erro QR/JSON"})
                else:
                    try:
                        partes = dados_lidos.split('-')
                        nova_linha.update({"Lote": partes[3].strip(), "Peso": float(partes[-1].strip()) / 1000.0})
                    except (ValueError, IndexError): nova_linha.update({"Lote": "erro QR/-", "Peso": "erro QR/-"})
            dados_processados.append(nova_linha)

        if not dados_processados: continue
        df_excel = pd.DataFrame(dados_processados)
        if 'Lote' in df_excel.columns: df_excel['Lote'] = df_excel['Lote'].fillna('').astype(str)
        output_filename = f"{os.path.splitext(os.path.basename(file_path))[0]}.xlsx"
        output_path = os.path.join(pasta_trabalho_temporaria, output_filename)
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_excel.to_excel(writer, index=False, sheet_name='Dados')
            worksheet = writer.sheets['Dados']
            for col_cells in worksheet.columns:
                max_len = max(len(str(cell.value)) for cell in col_cells if cell.value is not None)
                worksheet.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 2
            col_peso_letra = 'D' 
            for cell in worksheet[col_peso_letra]:
                if cell.row > 1: cell.number_format = '0.000'
    st.success("Etapa 1 concluída!")

    st.write("Etapa 2 de 2: Unificando todos os arquivos Excel em um relatório final...")
    arquivo_de_saida_final = os.path.join(pasta_trabalho_temporaria, "Inventario.xlsx")
    arquivos_excel_intermediarios = glob.glob(os.path.join(pasta_trabalho_temporaria, '*.xlsx'))
    if not arquivos_excel_intermediarios:
        st.error("Nenhum arquivo Excel intermediário foi gerado na Etapa 1.")
        return None
    
    lista_de_dataframes = []
    for arquivo in arquivos_excel_intermediarios:
        try:
            df = pd.read_excel(arquivo, dtype={'Lote': str})
            df['Localização'] = os.path.splitext(os.path.basename(arquivo))[0]
            lista_de_dataframes.append(df)
        except Exception as e: st.warning(f"Erro ao ler o arquivo intermediário {os.path.basename(arquivo)}. Erro: {e}")
    if not lista_de_dataframes:
        st.error("Nenhum dado foi lido dos arquivos intermediários. O arquivo final não será gerado.")
        return None

    df_final = pd.concat(lista_de_dataframes, ignore_index=True)
    with pd.ExcelWriter(arquivo_de_saida_final, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, sheet_name='Inventario_Unificado')
        worksheet = writer.sheets['Inventario_Unificado']
        for col_cells in worksheet.columns:
            max_len = max(len(str(cell.value)) for cell in col_cells if cell.value is not None)
            worksheet.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 2
    st.success("Etapa 2 concluída!")
    return arquivo_de_saida_final

# =====================================================================================
# INTERFACE DO STREAMLIT (UI)
# =====================================================================================

st.set_page_config(page_title="Conversor de Inventário", layout="wide")
st.title("Conversor de Inventário")

# --- INPUTS DO USUÁRIO ---
tipo_material = st.selectbox("Tipo de Material:", ["Produto Acabado", "Bobina"])
grupo_produto = st.text_input("Grupo de Produto:")
data_inventario = st.date_input("Data:", format="DD/MM/YYYY") # <-- MUDANÇA AQUI
uploaded_files = st.file_uploader(
    "Importar arquivos .csv",
    type="csv",
    accept_multiple_files=True
)

# --- BOTÃO E LÓGICA DE EXECUÇÃO ---
if st.button("Converter"):
    # Validações dos inputs
    if not uploaded_files:
        st.warning("Por favor, carregue pelo menos um arquivo .csv para converter.")
    elif not grupo_produto:
        st.warning("Por favor, preencha o campo 'Grupo de Produto'.")
    elif not data_inventario:
        st.warning("Por favor, selecione uma 'Data'.")
    else:
        with st.spinner("Aguarde... A conversão está em andamento."):
            try:
                with tempfile.TemporaryDirectory() as temp_dir:
                    pasta_csv_origem = os.path.join(temp_dir, "csv_original")
                    os.makedirs(pasta_csv_origem)

                    for uploaded_file in uploaded_files:
                        with open(os.path.join(pasta_csv_origem, uploaded_file.name), "wb") as f:
                            f.write(uploaded_file.getbuffer())

                    caminho_arquivo_final = None
                    if tipo_material == "Produto Acabado":
                        caminho_arquivo_final = processar_fluxo_produto_acabado(pasta_csv_origem, temp_dir)
                    elif tipo_material == "Bobina":
                        caminho_arquivo_final = processar_fluxo_bobina(pasta_csv_origem, temp_dir)

                    # --- DOWNLOAD DO ARQUIVO FINAL ---
                    if caminho_arquivo_final and os.path.exists(caminho_arquivo_final):
                        st.success("Conversão finalizada com sucesso!")
                        with open(caminho_arquivo_final, "rb") as file:
                            output_bytes = file.read()
                        
                        # Formata a data para dd-mm-aa para usar no nome do arquivo
                        data_formatada = data_inventario.strftime("%d-%m-%y")
                        nome_arquivo_download = f"Inventario {grupo_produto} {data_formatada}.xlsx"
                        
                        st.download_button(
                            label="Clique aqui para baixar o arquivo final",
                            data=output_bytes,
                            file_name=nome_arquivo_download,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    else:
                        st.error("Ocorreu um erro e o arquivo final não pôde ser gerado.")
            except Exception as e:
                st.error(f"Ocorreu um erro inesperado durante o processo: {e}")